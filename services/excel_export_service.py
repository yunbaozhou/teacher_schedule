import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO
from config import EXPORT_CONFIG, DEFAULT_COLORS
from services.color_service import get_course_color
from services.base_export_service import BaseExportService

class ExcelExportService(BaseExportService):
    """Excel export service"""
    
    def generate_excel(self, df, output, title):
        """Generate Excel file"""
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = title
            
            # Define border style with black color to match requirement
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Create a dictionary for quick course lookup
            course_dict = {}
            for _, course in df.iterrows():
                # Ensure weekday and period fields exist
                day = course.get('星期', '')
                period = course.get('节次', '')
                if day and period:
                    key = (day, period)
                    if key not in course_dict:
                        course_dict[key] = []
                    course_dict[key].append(course)
            
            # Set title
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']  # Get the top-left cell of the merged area
            title_cell.value = title
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.border = thin_border
            
            # Set headers
            headers = ['节次/星期', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.border = thin_border
            
            # Set morning/afternoon/evening study identifiers and courses
            time_periods = ['上午', '下午', '晚自习']
            # Each period start row: morning starts from row 3, afternoon from row 8, evening study from row 13
            period_starts = [3, 8, 13]
            
            # Store content for calculating column widths
            column_contents = {i: [] for i in range(1, 9)}
            
            # Keep track of assigned colors to ensure different courses have different colors
            assigned_colors = {}
            
            for i, (period, start_row) in enumerate(zip(time_periods, period_starts)):
                # Set period identifier (merge cells)
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
                period_cell = ws[f'A{start_row}']  # Get the top-left cell of the merged area
                period_cell.value = period
                period_cell.font = Font(bold=True)
                period_cell.alignment = Alignment(horizontal='center', vertical='center')
                period_cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
                period_cell.border = thin_border
                
                # Fill courses for this period (4 classes)
                for row_offset in range(4):
                    row = start_row + row_offset + 1  # Start filling courses from period identifier row + 1
                    period_num = row_offset + 1 + i * 4  # Period number: 1-4, 5-8, 9-12
                    # Set period
                    period_cell = ws.cell(row=row, column=1, value=f'第{period_num}节')
                    period_cell.alignment = Alignment(horizontal='center', vertical='center')
                    period_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    period_cell.border = thin_border
                    column_contents[1].append(period_cell.value)
                    
                    # Fill daily courses
                    days = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
                    for col, day in enumerate(days, 2):
                        # Convert "星期一" etc. to "周一" etc. to match data
                        day_short = day.replace('星期', '周')
                        key = (day_short, period_num)
                        if key in course_dict:
                            courses = course_dict[key]
                            # If there are multiple courses at the same time, display all courses
                            course_texts = []
                            for course in courses:
                                # Get course information, ensure fields exist
                                course_name = course.get('课程名称', '')
                                teacher = course.get('教师', '')
                                location = course.get('地点', '')
                                start_time = course.get('开始时间', '')
                                end_time = course.get('结束时间', '')
                                notes = course.get('备注', '')
                                
                                course_text = course_name
                                if teacher and teacher != '未指定':
                                    course_text += f"\n教师：{teacher}"
                                if location and location != '未指定':
                                    course_text += f"\n地点：{location}"
                                if start_time and end_time:
                                    course_text += f"\n时间：{start_time}~{end_time}"
                                if notes:
                                    course_text += f"\n备注：{notes}"
                                course_texts.append(course_text)
                            
                            cell_content = '\n'.join(course_texts)
                            cell = ws.cell(row=row, column=col, value=cell_content)
                            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                            cell.border = thin_border
                            column_contents[col].append(cell_content)
                            
                            # Get course color - ensure it matches frontend
                            first_course = courses[0]
                            course_name = first_course.get('课程名称', '')
                            user_selected_colors = first_course.get('user_selected_colors', {})
                            if course_name:
                                # Check if we've already assigned a color for this course name
                                if course_name in assigned_colors:
                                    # Use the already assigned color for consistency
                                    color_rgb = assigned_colors[course_name]
                                else:
                                    # Get color from color service
                                    color_rgb = get_course_color(course_name, user_selected_colors)
                                    
                                    # Ensure this color is not already used by another course
                                    used_colors = set(assigned_colors.values())
                                    color_index = 0
                                    # If the color is already used, find a different one
                                    while tuple(color_rgb) in used_colors and color_index < len(DEFAULT_COLORS):
                                        color_rgb = DEFAULT_COLORS[color_index]
                                        color_index += 1
                                    
                                    # Store the assigned color for this course name
                                    assigned_colors[course_name] = color_rgb
                                
                                color_hex = f"{color_rgb[0]:02X}{color_rgb[1]:02X}{color_rgb[2]:02X}"
                                fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                                cell.fill = fill
                        else:
                            # Even if no course, add empty content for width calculation
                            cell = ws.cell(row=row, column=col, value="")
                            cell.border = thin_border
                            column_contents[col].append("")
            
            # Set column widths based on content
            for col_idx in range(1, 9):
                # Calculate width based on content
                max_width = 12  # Minimum width
                for content in column_contents[col_idx]:
                    # Calculate max line width in multi-line content
                    lines = content.split('\n')
                    for line in lines:
                        # For Chinese characters, we need more width
                        line_width = 0
                        for char in line:
                            # Chinese characters take more space
                            if ord(char) > 127:  # Unicode value for non-ASCII characters
                                line_width += 2.5  # Increased width for Chinese characters
                            else:
                                line_width += 1.2  # Slightly more width for English characters
                        
                        max_width = max(max_width, line_width)
                
                # Apply adjustments - ensure time can be displayed in one line
                adjusted_width = min(max(max_width, 15), 50)  # Ensure minimum 15 for time display
                ws.column_dimensions[chr(64 + col_idx)].width = adjusted_width
            
            # Set row heights
            for row in range(1, 18):  # Update row range to fit new structure
                if row in [3, 8, 13]:  # Morning/afternoon/evening study rows
                    ws.row_dimensions[row].height = 25
                else:
                    ws.row_dimensions[row].height = 60  # Increased height for better text display
            
            # Save to output stream
            wb.save(output)
        except Exception as e:
            # Record detailed error information
            import traceback
            error_info = traceback.format_exc()
            print(f"Error generating Excel: {str(e)}")
            print(f"Detailed error information:\n{error_info}")
            raise e

    def create_excel_export(self, data):
        """Create Excel export"""
        df, title, _ = self.prepare_data(data)
        
        # Generate Excel file to memory
        output = BytesIO()
        self.generate_excel(df, output, title)
        output.seek(0)
        
        # Return file stream directly
        filename = f"{title}.xlsx"
        return output, filename